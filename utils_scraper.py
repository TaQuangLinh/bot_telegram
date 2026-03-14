
import re
import asyncio
import httpx
from io import BytesIO
from playwright.async_api import async_playwright
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
import random


def is_valid_mercari_url(url):
    patterns = [
        r"https://jp\.mercari\.com/item/m\d+",
        r"https://jp\.mercari\.com/en/item/m\d+",
        r"https://jp\.mercari\.com/shops/product/\w+",
        r"https://jp\.mercari\.com/en/shops/product/\w+"
    ]
    return any(re.match(p, url) for p in patterns)


async def scrape_one_page(page, url, customer):
    item_data = {
        "Tên khách": customer.lower(),
        "Tên sản phẩm": "Không tìm thấy / Lỗi",
        "Giá tiền (JPY)": "N/A",
        "Link": url,
        "Ảnh": ""
    }

    for attempt in range(3):
        try:
            # Tăng mức độ chờ đợi lên load thay vì domcontentloaded để ảnh kịp hiện link
            await page.goto(url, wait_until="load", timeout=60000)

            await page.wait_for_selector('h1', timeout=20000)
            item_data["Tên sản phẩm"] = await page.locator('h1').first.inner_text()

            # Lấy dữ liệu bằng evaluate
            product_data = await page.evaluate('''() => {
                const scripts = Array.from(document.querySelectorAll('script[type="application/ld+json"]'));
                for (let s of scripts) {
                    try {
                        let data = JSON.parse(s.innerText);
                        if (data['@type'] === 'Product') return data;
                        if (Array.isArray(data)) return data.find(d => d['@type'] === 'Product');
                    } catch(e) {}
                }
                return null;
            }''')

            if product_data:
                # Lấy giá
                price = product_data.get('offers', {}).get('price')
                if price: item_data["Giá tiền (JPY)"] = str(price)

                # Lấy ảnh từ JSON
                img_data = product_data.get('image', [])
                if img_data:
                    item_data["Ảnh"] = img_data[0] if isinstance(img_data, list) else img_data

            # --- VÉT CẠN NẾU THIẾU GIÁ HOẶC ẢNH ---

            # 1. Vét giá bằng Regex
            if item_data["Giá tiền (JPY)"] == "N/A":
                body_text = await page.locator('body').inner_text()
                price_match = re.search(r'¥\s*([\d,]+)', body_text)
                if price_match:
                    item_data["Giá tiền (JPY)"] = price_match.group(1).replace(',', '')

            # 2. Vét ảnh từ Meta hoặc thẻ IMG (Rất quan trọng)
            if not item_data["Ảnh"]:
                # Thử Meta og:image
                og_img = await page.get_attribute('meta[property="og:image"]', 'content')
                if og_img:
                    item_data["Ảnh"] = og_img
                else:
                    # Thử lấy ảnh đầu tiên trong carousel (thường có alt chứa tên SP)
                    img_src = await page.locator('figure img').first.get_attribute('src')
                    if img_src: item_data["Ảnh"] = img_src

            # Nếu đã có đủ cả Giá và Ảnh thì thoát Retry
            if item_data["Giá tiền (JPY)"] != "N/A" and item_data["Ảnh"]:
                break

            await asyncio.sleep(2)

        except Exception as e:
            if attempt == 2:
                print(f"Thất bại link {url}: {e}")
            else:
                await asyncio.sleep(3)

    return item_data

async def scrape_mercari_data(data_list):
    final_results = []
    # Chia nhỏ 10 link để tránh tràn RAM
    chunks = [data_list[i:i + 10] for i in range(0, len(data_list), 10)]

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True, args=["--no-sandbox", "--disable-dev-shm-usage"])
        # Đặt Default Timeout cho toàn bộ Context
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36"
        )
        context.set_default_timeout(60000)  # Ép mọi lệnh chờ tối đa 60s
        page = await context.new_page()

        for chunk in chunks:
            for item in chunk:
                res = await scrape_one_page(page, item['url'], item['customer'])
                final_results.append(res)
                await asyncio.sleep(random.uniform(1, 2))

            # Sau mỗi 10 link, đóng trang và mở trang mới để reset bộ nhớ
            await page.close()
            page = await context.new_page()

        await browser.close()
    return final_results


# create_excel_with_image giữ nguyên như cũ


def create_excel_with_image(results, output_file):
    df = pd.DataFrame(results)[["Tên sản phẩm", "Tên khách", "Link", "Giá tiền (JPY)", "Ảnh"]]
    df.to_excel(output_file, index=False, engine='openpyxl')
    wb = load_workbook(output_file)
    ws = wb.active
    ws.column_dimensions['E'].width = 25
    with httpx.Client(headers={'User-Agent': 'Mozilla/5.0'}, timeout=15) as client:
        for i, item in enumerate(results, start=2):
            ws.row_dimensions[i].height = 120
            if item.get("Ảnh") and str(item["Ảnh"]).startswith("http"):
                try:
                    res = client.get(item["Ảnh"])
                    if res.status_code == 200:
                        img = OpenpyxlImage(BytesIO(res.content))
                        img.width = img.height = 150
                        ws.add_image(img, f'E{i}')
                        ws[f'E{i}'] = ""
                except:
                    pass
    wb.save(output_file)